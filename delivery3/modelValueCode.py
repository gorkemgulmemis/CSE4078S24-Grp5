import openpyxl


def kiyasla(excelDosyasi):
    # Excel dosyasını aç
    workbook = openpyxl.load_workbook(excelDosyasi)
    sheet = workbook.active

    # Toplamda kaç tane 1 ve 0 olduğunu sayacak değişkenler
    bir_sayaci = 0
    sifir_sayaci = 0

    # İlk 251 satıra bak
    for satir in range(1, 252):
        # D ve F sütunlarını al, büyük-küçük harf duyarlılığını dikkate almadan
        deger_D = sheet.cell(row=satir, column=4).value.lower()
        deger_F = sheet.cell(row=satir, column=6).value.lower()

        # Eğer D ve F sütunları aynı ise
        if deger_D == deger_F:
            # G, H, I, J ve K sütunlarına 1 yaz
            for sutun in range(7, 12):
                sheet.cell(row=satir, column=sutun).value = 1
            bir_sayaci += 1
        else:
            # G, H, I, J ve K sütunlarına 0 yaz
            for sutun in range(7, 12):
                sheet.cell(row=satir, column=sutun).value = 0
            sifir_sayaci += 1

    # Excel dosyasını kaydet
    workbook.save(excelDosyasi)

    # Toplamda kaç tane 1 ve 0 olduğunu ekrana yazdır
    print("Toplamda {} satırda 1 yazıldı.".format(bir_sayaci))
    print("Toplamda {} satırda 0 yazıldı.".format(sifir_sayaci))


# Kodu çalıştır
kiyasla("NLP Tablo.xlsx")
