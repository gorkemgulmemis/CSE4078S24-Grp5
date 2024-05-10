import openpyxl

def kiyasla(excelDosyasi):
    # Excel dosyasını aç
    workbook = openpyxl.load_workbook(excelDosyasi)
    sheet = workbook.active

    # Toplamda kaç tane aynı var sayacı
    ayni_sayaci = 0

    # İlk 251 satıra bak
    for satir in range(1, 252):
        # T ve U sütunlarını al
        kelime_T = sheet.cell(row=satir, column=5).value
        kelime_U = sheet.cell(row=satir, column=4).value

        # Her iki hücre de doluysa ve büyük-küçük harf duyarlılığı olmadan aynıysa
        if kelime_T and kelime_U and kelime_T.lower() == kelime_U.lower():
            ayni_sayaci += 1

    # Toplamda kaç tane aynı olduğunu ekrana yazdır
    print("Toplamda {} satırda T ve U sütunları aynı kelimeleri içeriyor.".format(ayni_sayaci))

# Kodu çalıştır
kiyasla("NLP Tablo.xlsx")
